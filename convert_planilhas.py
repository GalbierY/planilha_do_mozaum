"""Convert the social assistance Word tables into Excel tabs for 2025 base and reporting."""
from itertools import zip_longest
from pathlib import Path
import textwrap
import unicodedata

import pandas as pd
from docx import Document
from openpyxl import load_workbook
from openpyxl.styles import Alignment

DOC_DIR = Path(__file__).resolve().parent / "Planilhas"
OUTPUT_PATH = DOC_DIR / "AssistenteSocial.xlsx"

HEADER_MAP = {
    "ALUNO": "Aluno",
    "NOME": "Aluno",
    "IDADE": "Idade",
    "ESCOLA": "Escola",
    "DATA DE NASCI": "Data de nascimento",
    "DATA DE NASCIMENTO": "Data de nascimento",
    "DATA EM QUE RECEBI O ENC": "Data recebimento",
    "DATA EM QUE RECEBI O ENCAMINHAMENTO": "Data recebimento",
    "ATEND REALIZADO": "Atendimento realizado",
    "VD": "VD",
    "GRAU DE RISCO": "Grau de risco",
    "BAIRRO": "Bairro",
    "ANO": "Ano escolar",
    "ACOMPANHAMENTO ESCOLAR": "Acompanhamento escolar",
}
BASE_COLUMNS = [
    "Aluno",
    "Idade",
    "Escola",
    "Data de nascimento",
    "Atendimento realizado",
    "VD",
]

REPORT_COLUMNS = [
    "Fonte",
    "Aluno",
    "Idade",
    "Escola",
    "Data de nascimento",
    "Atendimento realizado",
    "VD",
    "Data recebimento",
    "Grau de risco",
    "Bairro",
    "Ano escolar",
    "Acompanhamento escolar",
]


WRAP_LIMITS = {
    "Atendimento realizado": 90,
    "Acompanhamento escolar": 80,
    "VD": 60,
}


def to_float(series: pd.Series) -> pd.Series:
    extracted = series.astype(str).str.extract(r"(\d+(?:[.,]\d+)?)")
    return pd.to_numeric(extracted[0].str.replace(",", "."), errors="coerce")


def build_stats_sheet(base: pd.DataFrame, report: pd.DataFrame) -> pd.DataFrame:
    stats = []
    base_age = to_float(base["Idade"])
    base_top = base["Escola"].value_counts().head(3)
    stats.append({"Category": "Base 2025", "Metric": "Total registros", "Value": len(base)})
    stats.append({"Category": "Base 2025", "Metric": "Média de idade", "Value": f"{base_age.mean():.2f}"})
    stats.append({"Category": "Base 2025", "Metric": "Idade mínima", "Value": f"{base_age.min():.0f}"})
    stats.append({"Category": "Base 2025", "Metric": "Idade máxima", "Value": f"{base_age.max():.0f}"})
    stats.append({"Category": "Base 2025", "Metric": "Escolas distintas", "Value": base["Escola"].nunique(dropna=True)})
    top3 = "; ".join(f"{idx} ({cnt})" for idx, cnt in base_top.items())
    stats.append({"Category": "Base 2025", "Metric": "Top 3 escolas", "Value": top3 or "—"})

    report_age = to_float(report["Idade"])
    stats.append({"Category": "Relatórios", "Metric": "Total linhas", "Value": len(report)})
    fonte_counts = report["Fonte"].value_counts()
    for fonte, count in fonte_counts.items():
        stats.append({"Category": "Relatórios", "Metric": f"Linhas {fonte}", "Value": count})
    stats.append({"Category": "Relatórios", "Metric": "Média de idade", "Value": f"{report_age.mean():.2f}"})
    vd_count = report["VD"].astype(bool).sum()
    stats.append({"Category": "Relatórios", "Metric": "VD com texto", "Value": vd_count})

    cases_mask = report["Fonte"].str.lower().str.contains("casos")
    cases = report[cases_mask]
    stats.append({"Category": "Casos judiciais", "Metric": "Total casos", "Value": len(cases)})
    acomp = cases["Acompanhamento escolar"].fillna("").replace("", "—")
    acomp_top = acomp.value_counts().head(5)
    stats.append({"Category": "Casos judiciais", "Metric": "Principais observações", "Value": "; ".join(f"{idx} ({cnt})" for idx, cnt in acomp_top.items())})

    stats.append({"Category": "Resumo geral", "Metric": "Alunos 2025", "Value": len(base)})
    stats.append({"Category": "Resumo geral", "Metric": "Alunos 2026", "Value": fonte_counts.get("2026", 0)})
    stats.append({"Category": "Resumo geral", "Metric": "Casos judiciais", "Value": len(cases)})
    stats.append({"Category": "Resumo geral", "Metric": "Escolas no relatório", "Value": report["Escola"].nunique(dropna=True)})
    return pd.DataFrame(stats)


def normalize_text(text: str) -> str:
    if not text:
        return ""
    cleaned = " ".join(str(text).split())
    return cleaned


def strip_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text or "")
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def canonical_header(text: str) -> str:
    cleaned = strip_accents(text).replace(".", "").strip().upper()
    cleaned = " ".join(cleaned.split())
    if not cleaned:
        return ""
    return HEADER_MAP.get(cleaned, cleaned.title())


def normalized_row_text(text: str) -> str:
    cleaned = strip_accents(text).replace("\n", " ").replace("\r", " ")
    cleaned = " ".join(cleaned.split())
    return cleaned.upper()


def apply_wrap_columns(frame: pd.DataFrame) -> None:
    for column, limit in WRAP_LIMITS.items():
        if column not in frame.columns:
            continue
        frame[column] = frame[column].astype(str).apply(lambda v: wrap_long_text(v, limit))


def wrap_long_text(value: str, limit: int) -> str:
    if not isinstance(value, str):
        return value
    if not value.strip():
        return value
    if "\n" in value or len(value) <= limit:
        return value
    lines = textwrap.wrap(value, width=limit, break_long_words=False, replace_whitespace=False)
    return "\n".join(lines)


def table_to_dataframe(table, header_marker: str) -> pd.DataFrame | None:
    marker = normalized_row_text(header_marker)
    header_row_index = None
    for idx, row in enumerate(table.rows):
        combined = " ".join(cell.text for cell in row.cells)
        if marker in normalized_row_text(combined):
            header_row_index = idx
            break
    if header_row_index is None:
        return None

    header_cells = table.rows[header_row_index].cells
    headers = [canonical_header(cell.text) for cell in header_cells]
    data_rows = []
    for row in table.rows[header_row_index + 1:]:
        values = [normalize_text(cell.text) for cell in row.cells]
        if not any(values):
            continue
        record = {}
        for header, value in zip_longest(headers, values, fillvalue=""):
            if not header:
                continue
            record[header] = value
        if record:
            data_rows.append(record)
    if not data_rows:
        return None
    return pd.DataFrame(data_rows)


def read_docx(path: Path, header_marker: str) -> pd.DataFrame:
    document = Document(path)
    frames = []
    for table in document.tables:
        frame = table_to_dataframe(table, header_marker)
        if frame is not None:
            frames.append(frame)
    if not frames:
        raise ValueError(f"No tables with header '{header_marker}' found in {path.name}")
    return pd.concat(frames, ignore_index=True)


def build_report_dataframe() -> pd.DataFrame:
    path_2026 = DOC_DIR / "PLANILHA ALUNOS ASSISTENTE SOCIAL 2026.docx"
    df_2026 = read_docx(path_2026, header_marker="ALUNO")
    body_columns = REPORT_COLUMNS[1:]
    report_2026 = df_2026.reindex(columns=body_columns, fill_value="").copy()
    report_2026.insert(0, "Fonte", "2026")
    return report_2026


def build_cases_dataframe() -> pd.DataFrame:
    path_casos = DOC_DIR / "CASOS_JUDICIAIS.docx"
    df_casos = read_docx(path_casos, header_marker="NOME")
    body_columns = REPORT_COLUMNS[1:]
    cases = pd.DataFrame(index=df_casos.index, columns=body_columns)
    cases = cases.fillna("")
    if "Aluno" in df_casos:
        cases["Aluno"] = df_casos["Aluno"].fillna("")
    if "Ano escolar" in df_casos:
        cases["Ano escolar"] = df_casos["Ano escolar"].fillna("")
    if "Acompanhamento escolar" in df_casos:
        cases["Acompanhamento escolar"] = df_casos["Acompanhamento escolar"].fillna("")
    cases.insert(0, "Fonte", "Casos judiciais")
    return cases


def main() -> None:
    source_2025 = DOC_DIR / "PLANILHA ALUNOS ASSISTENTE SOCIAL 2025.docx"
    df_2025 = read_docx(source_2025, header_marker="ALUNO")
    base_sheet = df_2025.reindex(columns=BASE_COLUMNS, fill_value="")
    report_frames = [build_report_dataframe(), build_cases_dataframe()]
    report_sheet = pd.concat(report_frames, ignore_index=True)
    apply_wrap_columns(base_sheet)
    apply_wrap_columns(report_sheet)
    stats_sheet = build_stats_sheet(base_sheet, report_sheet)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        base_sheet.to_excel(writer, sheet_name="Base2025", index=False)
        report_sheet.to_excel(writer, sheet_name="Relatorios", index=False)
        stats_sheet.to_excel(writer, sheet_name="Estatisticas", index=False)

    adjust_excel_layout(OUTPUT_PATH, ["Base2025", "Relatorios", "Estatisticas"])
    print(f"Wrote Excel workbook to {OUTPUT_PATH}")


def adjust_excel_layout(path: Path, sheet_names: list[str]) -> None:
    workbook = load_workbook(path)
    wrap = Alignment(wrap_text=True, vertical="top")
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = wrap
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value is None:
                    continue
                cell_value = str(cell.value)
                max_length = max(max_length, len(cell_value))
            adjusted_width = max(10, max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width
    workbook.save(path)


if __name__ == "__main__":
    main()
